"""
ng_otvety.py — адаптирован для TDM.

ИЗМЕНЕНО:
  - parcing_data() — убран Telegram-контекст, добавлен on_error callback
  - async-обёртка сохранена для обратной совместимости

БЕЗ ИЗМЕНЕНИЙ (скопируй из оригинала):
  - excluded_dates (список дат)
  - choosing_day()
  - choosing_time_NG()
  - process_ng_prosroki_file()
  - personalizating_table_osn()
  - personalizating_table_prosrok()
  - personalizating_table_eight_day()
  - personalizating_table_seven_day()
  - personalizating_table_six_day()
  - personalizating_table_five_day()
  - add_run_delete_and_save_files()
"""

import os
import time
from datetime import datetime, timedelta

import pandas as pd
from dotenv import load_dotenv

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

# --- все остальные импорты из оригинала (openpyxl, win32com и т.д.) ---
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
import win32com.client
from functools import reduce
import pythoncom

load_dotenv()
login_NG = os.getenv("login_NG")
password_NG = os.getenv("password_NG")

home_dir = os.path.expanduser("~")
directory = os.path.join(home_dir, "Downloads")


excluded_dates = [
    # Майские праздники (1-3 мая и 9-11 мая)
    "01.05.2026", "02.05.2026", "03.05.2026",
    "09.05.2026", "10.05.2026", "11.05.2026",
    # Май (добавлены только неохваченные)
    "16.05.2026", "17.05.2026", "23.05.2026", "24.05.2026", "30.05.2026", "31.05.2026",
    # Июнь (добавлены только неохваченные)
    "06.06.2026", "07.06.2026", "20.06.2026", "21.06.2026", "27.06.2026", "28.06.2026", "12.06.2026", "13.06.2026", "14.06.2026",
    # Июль
    "04.07.2026", "05.07.2026", "11.07.2026", "12.07.2026", "18.07.2026", "19.07.2026", "25.07.2026", "26.07.2026",
    # Август
    "01.08.2026", "02.08.2026", "08.08.2026", "09.08.2026", "15.08.2026", "16.08.2026", "22.08.2026", "23.08.2026",
    "29.08.2026", "30.08.2026",
    # Сентябрь
    "05.09.2026", "06.09.2026", "12.09.2026", "13.09.2026", "19.09.2026", "20.09.2026", "26.09.2026", "27.09.2026",
    # Октябрь (добавлены только неохваченные)
    "10.10.2026", "11.10.2026", "17.10.2026", "18.10.2026", "24.10.2026", "25.10.2026",
    # Ноябрь (добавлены только неохваченные)
    "14.11.2026", "15.11.2026", "21.11.2026", "22.11.2026", "28.11.2026", "29.11.2026",
    # Декабрь
    "05.12.2026", "06.12.2026", "12.12.2026", "13.12.2026", "19.12.2026", "20.12.2026", "26.12.2026", "27.12.2026",
    # Канун Нового 2027 года
    "31.12.2026"
]


# ──────────────────────────────────────────────
# ИЗМЕНЁННАЯ ФУНКЦИЯ: parcing_data
# ──────────────────────────────────────────────

def parcing_data_sync(on_error=None) -> bool:
    """
    Синхронный парсинг «Ответов в работе» (НГ).
    on_error(text) — опциональный callback для уведомления об ошибке.
    """
    chrome_install = ChromeDriverManager().install()
    folder = os.path.dirname(chrome_install)
    chromedriver_path = os.path.join(folder, "chromedriver.exe")
    driver = webdriver.Chrome(service=ChromeService(chromedriver_path))

    try:
        driver.get("https://gorod.mos.ru/api/service/auth/auth")

        driver.find_element(By.XPATH, '//input[@placeholder="Логин *"]').send_keys(login_NG)
        driver.find_element(By.XPATH, '//input[@placeholder="Пароль*"]').send_keys(password_NG)
        driver.find_element(
            By.XPATH, "/html/body/div[1]/div/div/main/div/div/div/div[2]/form[1]/button"
        ).click()

        WebDriverWait(driver, 200).until(EC.presence_of_element_located((By.XPATH,
                                                                         '//div[@class="dashboard__block-link"]'
                                                                         '//div[@class="button-big link"]'
                                                                         '//div[@class="dashboard-container__links-title" and contains(text(), "Аналитика")]'
                                                                         )))

        driver.get("https://gorod.mos.ru/admin/ker/olap/report/155")
        time.sleep(7)

        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH,
                                                                        "/html/body/div[3]/div/div[2]/div/div/div/div/form/footer/button[3]/span[2]/span"
                                                                        )))
        driver.find_element(By.XPATH,
                            "/html/body/div[3]/div/div[2]/div/div/div/div/form/footer/button[3]/span[2]/span"
                            ).click()
        time.sleep(1)

        driver.find_element(By.XPATH,
                            "//button[contains(@class, 'bg-primary')]//span[text()='Экспорт']"
                            ).click()
        time.sleep(1)

        driver.get("https://gorod.mos.ru/admin/ker/olap/downloads")
        WebDriverWait(driver, 1500).until(EC.presence_of_element_located((By.XPATH,
                                                                          "/html/body/div[1]/div/div[2]/main/div/div[1]/div/div[2]"
                                                                          "/div[1]/table/tbody/tr[1]/td[5]/div/i"
                                                                          )))
        driver.find_element(By.XPATH,
                            "/html/body/div[1]/div/div[2]/main/div/div[1]/div/div[2]"
                            "/div[1]/table/tbody/tr[1]/td[5]/div/i"
                            ).click()
        time.sleep(15)
        return True

    except Exception as e:
        msg = f"❌ Ошибка при выгрузке Ответы в работе (НГ): {e}"
        print(msg)
        if on_error:
            on_error(msg)
        return False
    finally:
        driver.quit()


# Async-обёртка — сохранена для обратной совместимости
async def parcing_data(context=None, chat_id=None):
    import asyncio
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, parcing_data_sync)


def choosing_day(excluded_date):
    today = datetime.now().date()
    user_input = today
    days_count = 8
    excluded_date = [datetime.strptime(date, "%d.%m.%Y").date() for date in
                     excluded_date]  # делаем даты удобными для прочтения, к одному формату
    # основной цикл для нахождения даты
    while days_count != 0:
        if user_input in excluded_date:
            user_input += timedelta(days=1)
        else:
            user_input += timedelta(days=1)
            days_count -= 1
    print(user_input)
    return user_input


def choosing_time_NG():
    timenow = pd.Timestamp(datetime.now()).strftime('%H-%M')
    return timenow


def process_ng_prosroki_file(timenow, filepath, excluded_dates):
    user_input = choosing_day(excluded_dates)
    df = pd.read_excel(filepath)

    df['Регламентный срок у сообщения (Портал)'] = df['Регламентный срок у сообщения (Портал)'].apply(
        lambda x: x.replace(second=0))
    df = df[df['Регламентный срок у сообщения (Портал)'] <= pd.to_datetime(user_input)]
    today = datetime.now()

    responsible_mapping = {
        'ГБУ «Автомобильные дороги ЮВАО»': 'АВД ЮВАО',
        'ГБУ Жилищник Выхино района Выхино-Жулебино города Москвы': 'Выхино-Жулебино',
        'Управа Выхино-Жулебино': 'Выхино-Жулебино',
        'ГБУ Жилищник Нижегородского района города Москвы': 'Нижегородский',
        'Управа Нижегородский': 'Нижегородский',
        'ГБУ Жилищник района Капотня города Москвы': 'Капотня',
        'Управа Капотня': 'Капотня',
        'ГБУ Жилищник района Кузьминки города Москвы': 'Кузьминки',
        'Управа Кузьминки': 'Кузьминки',
        'ГБУ Жилищник района Лефортово города Москвы': 'Лефортово',
        'Управа Лефортово': 'Лефортово',
        'ГБУ Жилищник района Люблино города Москвы': 'Люблино',
        'Управа Люблино': 'Люблино',
        'ГБУ Жилищник района Марьино города Москвы': 'Марьино',
        'Управа Марьино': 'Марьино',
        'ГБУ Жилищник района Некрасовка города Москвы': 'Некрасовка',
        'Управа Некрасовка': 'Некрасовка',
        'ГБУ Жилищник района Печатники города Москвы': 'Печатники',
        'Управа Печатники': 'Печатники',
        'ГБУ Жилищник района Текстильщики города Москвы': 'Текстильщики',
        'Управа Текстильщики': 'Текстильщики',
        'ГБУ Жилищник Рязанского района города Москвы': 'Рязанский',
        'Управа Рязанский': 'Рязанский',
        'ГБУ Жилищник Южнопортового района города Москвы': 'Южнопортовый',
        'Управа Южнопортовый': 'Южнопортовый'
    }

    districts_index = ['АВД ЮВАО', 'Выхино-Жулебино', 'Капотня', 'Кузьминки', 'Лефортово',
                       'Люблино', 'Марьино', 'Некрасовка', 'Нижегородский', 'Печатники',
                       'Рязанский', 'Текстильщики', 'Южнопортовый']

    df['Район'] = df['Ответственный ОИВ первого уровня'].map(responsible_mapping)

    valid_organizations = list(responsible_mapping.keys())
    valid_organizations.append('Префектура Юго-Восточного округа')
    df = df[df['Ответственный за подготовку ответа'].isin(valid_organizations)]

    # просрочки ЛК префекта
    condition = (df['Ответственный за подготовку ответа'].str.contains('Префектура Юго-Восточного округа')) & (
            df['Регламентный срок у сообщения (Портал)'] < today)
    prefect = df[condition].copy()

    pivot_prefect = pd.pivot_table(prefect, values='Номер заявки', index='Район', aggfunc='count')
    pivot_prefect = pivot_prefect.rename(columns={'Номер заявки': 'Кабинет префекта просрочки'})
    if pivot_prefect.empty:
        pivot_prefect = pd.DataFrame(index=districts_index, columns=['Кабинет префекта просрочки'])

    df = df[~df['Ответственный за подготовку ответа'].str.contains('Префектура Юго-Восточного округа')]

    excluded_dates_with_time = [
        datetime.strptime(date_str, "%d.%m.%Y").replace(hour=23, minute=59, second=0)
        for date_str in excluded_dates
    ]
    excluded_dates_dt = pd.to_datetime(excluded_dates_with_time)
    excluded_date = [datetime.strptime(date, "%d.%m.%Y").date() for date in excluded_dates]

    main_df = df.copy()

    def change_status(df):
        df = df.copy()
        repl = {
            "Готовится ответ": "Готовится ответ (ОИВ взял доп. срок)",
            "На доработке": "На доработке (Город вернул)",
            "На модерации": "На модерации (Проверка города)",
            "На утверждении": "На утверждении (У куратора)",
            "Нет ответа": "Нет ответа (ОИВ не дал ответ)",
        }
        df.loc[:, "Статус подготовки ответа на сообщение"] = df[
            "Статус подготовки ответа на сообщение"].replace(repl)
        return df

    def table_is_none(date, number):
        df_empty = pd.DataFrame(index=districts_index,
                                columns=[f'{number} день ({date.strftime("%d.%m")})']).fillna(0)
        print(f"{number}-й день пустой")
        return df_empty

    def crearing_day_in_svod(df, date, number):
        new_date = date + timedelta(days=1)
        while new_date in excluded_date:
            new_date += timedelta(days=1)
        if df.empty:
            return table_is_none(new_date, number), new_date
        df_date = change_status(df[df['Регламентный срок у сообщения (Портал)'].dt.date == new_date])
        pivot_date_for_svod = pd.pivot_table(df_date, values='Номер заявки', index='Район', aggfunc='count')
        new_name = f'{number} день ({new_date.strftime("%d.%m")})'
        if not pivot_date_for_svod.empty:
            pivot_date_for_svod.rename(columns={pivot_date_for_svod.columns[-1]: new_name}, inplace=True)
        else:
            pivot_date_for_svod = table_is_none(new_date, number)
        return pivot_date_for_svod, new_date

    # ---- 8-й день ----
    today_date = datetime.now().date()
    day_8 = today_date
    while day_8 in excluded_date:
        day_8 += timedelta(days=1)
    df_date_8 = change_status(main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date == day_8]) \
        if not main_df.empty else pd.DataFrame(columns=main_df.columns)
    pivot8_dlya_svoda = pd.pivot_table(df_date_8, values='Номер заявки', index='Район', aggfunc='count')
    if not pivot8_dlya_svoda.empty:
        pivot8_dlya_svoda.rename(columns={pivot8_dlya_svoda.columns[-1]: f'8 день ({day_8.strftime("%d.%m")})'},
                                 inplace=True)
    else:
        pivot8_dlya_svoda = table_is_none(day_8, 8)
    pivot_8 = pd.pivot_table(df_date_8, values='Номер заявки', index='Район',
                             columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)
    if not pivot_8.empty:
        pivot_8.rename(columns={pivot_8.columns[-1]: 'Всего'}, inplace=True)
        pivot_8.rename(index={pivot_8.index[-1]: 'Всего'}, inplace=True)

    # ---- 7-й день ----
    day_7 = day_8 + timedelta(days=1)
    while day_7 in excluded_date:
        day_7 += timedelta(days=1)
    df_date_7 = change_status(main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date == day_7]) \
        if not main_df.empty else pd.DataFrame(columns=main_df.columns)
    pivot7_dlya_svoda = pd.pivot_table(df_date_7, values='Номер заявки', index='Район', aggfunc='count')
    if not pivot7_dlya_svoda.empty:
        pivot7_dlya_svoda.rename(columns={pivot7_dlya_svoda.columns[-1]: f'7 день ({day_7.strftime("%d.%m")})'},
                                 inplace=True)
    else:
        pivot7_dlya_svoda = table_is_none(day_7, 7)
    pivot_7 = pd.pivot_table(df_date_7, values='Номер заявки', index='Район',
                             columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)
    if not pivot_7.empty:
        pivot_7.rename(columns={pivot_7.columns[-1]: 'Всего'}, inplace=True)
        pivot_7.rename(index={pivot_7.index[-1]: 'Всего'}, inplace=True)

    # ---- 6-й день ----
    day_6 = day_7 + timedelta(days=1)
    while day_6 in excluded_date:
        day_6 += timedelta(days=1)
    df_date_6 = change_status(main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date == day_6]) \
        if not main_df.empty else pd.DataFrame(columns=main_df.columns)
    pivot6_dlya_svoda = pd.pivot_table(df_date_6, values='Номер заявки', index='Район', aggfunc='count')
    if not pivot6_dlya_svoda.empty:
        pivot6_dlya_svoda.rename(columns={pivot6_dlya_svoda.columns[-1]: f'6 день ({day_6.strftime("%d.%m")})'},
                                 inplace=True)
    else:
        pivot6_dlya_svoda = table_is_none(day_6, 6)
    pivot_6 = pd.pivot_table(df_date_6, values='Номер заявки', index='Район',
                             columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)
    if not pivot_6.empty:
        pivot_6.rename(columns={pivot_6.columns[-1]: 'Всего'}, inplace=True)
        pivot_6.rename(index={pivot_6.index[-1]: 'Всего'}, inplace=True)

    # ---- 5-й день ----
    day_5 = day_6 + timedelta(days=1)
    while day_5 in excluded_date:
        day_5 += timedelta(days=1)
    df_date_5 = change_status(main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date == day_5]) \
        if not main_df.empty else pd.DataFrame(columns=main_df.columns)
    pivot5_dlya_svoda = pd.pivot_table(df_date_5, values='Номер заявки', index='Район', aggfunc='count')
    if not pivot5_dlya_svoda.empty:
        pivot5_dlya_svoda.rename(columns={pivot5_dlya_svoda.columns[-1]: f'5 день ({day_5.strftime("%d.%m")})'},
                                 inplace=True)
    else:
        pivot5_dlya_svoda = table_is_none(day_5, 5)
    pivot_5 = pd.pivot_table(df_date_5, values='Номер заявки', index='Район',
                             columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)
    if not pivot_5.empty:
        pivot_5.rename(columns={pivot_5.columns[-1]: 'Всего'}, inplace=True)
        pivot_5.rename(index={pivot_5.index[-1]: 'Всего'}, inplace=True)

    # ---- дни 4..1 ----
    pivot4_dlya_svoda, date4 = crearing_day_in_svod(main_df, day_5, 4)
    pivot3_dlya_svoda, date3 = crearing_day_in_svod(main_df, date4, 3)
    pivot2_dlya_svoda, date2 = crearing_day_in_svod(main_df, date3, 2)
    pivot1_dlya_svoda, date1 = crearing_day_in_svod(main_df, date2, 1)

    # ---- ПРОСРОЧКИ ----
    prosrok = main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date < today_date] \
        if not main_df.empty else pd.DataFrame(columns=main_df.columns)
    prosrok_for_svod = pd.pivot_table(prosrok, values='Номер заявки', index='Район', aggfunc='count')
    prosrok_for_svod = prosrok_for_svod.rename(columns={'Номер заявки': 'Просрочки'})
    if prosrok_for_svod.empty:
        prosrok_for_svod = pd.DataFrame(index=districts_index, columns=['Просрочки']).fillna(0)

    df_prosrok = change_status(prosrok)
    pivot_prosrok = pd.pivot_table(df_prosrok, values='Номер заявки', index='Район',
                                   columns="Статус подготовки ответа на сообщение", aggfunc='count',
                                   margins=True) if not df_prosrok.empty else pd.DataFrame()
    if not pivot_prosrok.empty:
        pivot_prosrok.rename(columns={pivot_prosrok.columns[-1]: 'Всего'}, inplace=True)
        pivot_prosrok.rename(index={pivot_prosrok.index[-1]: 'Всего'}, inplace=True)

    # ====== ВЫХОДНЫЕ / ПРАЗДНИЧНЫЕ ДНИ ВНУТРИ ОКНА ======
    def build_holiday_block(block_dates):
        if len(block_dates) == 1:
            label = f"Вых. дни ({block_dates[0].strftime('%d.%m')})"
        else:
            label = f"Вых. дни ({block_dates[0].strftime('%d.%m')}-{block_dates[-1].strftime('%d.%m')})"
        sub = change_status(
            main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date.isin(block_dates)]) \
            if not main_df.empty else pd.DataFrame(columns=main_df.columns)
        piv = pd.pivot_table(sub, values='Номер заявки', index='Район', aggfunc='count')
        if not piv.empty:
            piv.rename(columns={piv.columns[-1]: label}, inplace=True)
        else:
            piv = pd.DataFrame(index=districts_index, columns=[label]).fillna(0)
        return piv, label

    day_pivots = [
        (day_8, pivot8_dlya_svoda), (day_7, pivot7_dlya_svoda),
        (day_6, pivot6_dlya_svoda), (day_5, pivot5_dlya_svoda),
        (date4, pivot4_dlya_svoda), (date3, pivot3_dlya_svoda),
        (date2, pivot2_dlya_svoda), (date1, pivot1_dlya_svoda),
    ]

    # Собираем столбцы по порядку дат; между рабочими днями вставляем блоки нерабочих
    all_weekend_dates = []        # все нерабочие даты окна (для листа по статусам)
    ordered_pivots = [prosrok_for_svod]
    urgent_weekend_labels = []    # блоки внутри диапазона 8–5 дня

    # Ведущий блок: если СЕГОДНЯ выходной/праздник — дни от сегодня до 8-го дня
    leading_gap = [today_date + timedelta(days=k) for k in range((day_8 - today_date).days)]
    if leading_gap:
        all_weekend_dates.extend(leading_gap)
        block_piv, block_label = build_holiday_block(leading_gap)
        ordered_pivots.append(block_piv)
        if leading_gap[-1] < day_5:
            urgent_weekend_labels.append(block_label)

    for i, (d, piv) in enumerate(day_pivots):
        if i > 0:
            prev_d = day_pivots[i - 1][0]
            gap = [prev_d + timedelta(days=k) for k in range(1, (d - prev_d).days)]  # пропущенные нерабочие даты
            if gap:
                all_weekend_dates.extend(gap)
                block_piv, block_label = build_holiday_block(gap)
                ordered_pivots.append(block_piv)
                if gap[-1] < day_5:           # блок заканчивается до 5-го дня → срочный
                    urgent_weekend_labels.append(block_label)
        ordered_pivots.append(piv)

    merged_df = reduce(lambda left, right: pd.merge(left, right, left_index=True, right_index=True, how='outer'),
                       ordered_pivots)

    # ---- статусная разбивка по ВСЕМ выходным/праздничным дням окна (отдельный лист) ----
    if all_weekend_dates and not main_df.empty:
        df_weekend_all = change_status(
            main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date.isin(all_weekend_dates)])
    else:
        df_weekend_all = pd.DataFrame(columns=main_df.columns)
    pivot_weekend = pd.pivot_table(df_weekend_all, values='Номер заявки', index='Район',
                                   columns="Статус подготовки ответа на сообщение", aggfunc='count',
                                   margins=True) if not df_weekend_all.empty else pd.DataFrame()
    if not pivot_weekend.empty:
        pivot_weekend.rename(columns={pivot_weekend.columns[-1]: 'Всего'}, inplace=True)
        pivot_weekend.rename(index={pivot_weekend.index[-1]: 'Всего'}, inplace=True)

    # ---- датафреймы для листов "Выходные" (детально) и "Префект просрок" ----
    holidays_df = main_df[main_df['Регламентный срок у сообщения (Портал)'].isin(excluded_dates_dt)]
    main_df = main_df[~main_df['Регламентный срок у сообщения (Портал)'].isin(excluded_dates_dt)].sort_values(
        by='Регламентный срок у сообщения (Портал)')

    # ====== СВОД ======
    merged_table = pd.merge(pivot_prefect, merged_df, left_index=True, right_index=True, how='outer').fillna(0)
    all_in_work = pd.DataFrame({'Всего в работе': merged_table.sum(axis=1)}).fillna(0)

    # "Всего срочных" считаем по названиям колонок (8–5 день + просрочки + срочные выходные)
    name_8 = f'8 день ({day_8.strftime("%d.%m")})'
    name_7 = f'7 день ({day_7.strftime("%d.%m")})'
    name_6 = f'6 день ({day_6.strftime("%d.%m")})'
    name_5 = f'5 день ({day_5.strftime("%d.%m")})'
    urgent_cols = ['Кабинет префекта просрочки', 'Просрочки', name_8, name_7, name_6, name_5] + urgent_weekend_labels
    urgent_cols = [c for c in urgent_cols if c in merged_table.columns]
    all_urgent = pd.DataFrame({'Всего срочных': merged_table[urgent_cols].sum(axis=1)}).fillna(0)

    final_svod = pd.merge(all_in_work, pivot_prefect, left_index=True, right_index=True, how='outer').fillna(0)
    final_svod = pd.merge(final_svod, all_urgent, left_index=True, right_index=True, how='outer').fillna(0)
    final_svod = pd.merge(final_svod, merged_df, left_index=True, right_index=True, how='outer').fillna(0)
    final_svod = final_svod.sort_values(
        by=['Всего срочных', 'Всего в работе'],
        ascending=[False, False]
    )

    totals_row = final_svod.sum(axis=0)
    totals_row.name = 'Итог по округу'
    df_with_totals = pd.concat([final_svod, pd.DataFrame(totals_row).T])
    df_with_totals.index.name = 'Ответственный за подготовку ответа'

    # ---- сохранение (новый лист "Выходные статусы" идёт сразу после "5-й день") ----
    processed_file_path = os.path.join(directory,
                                       f"Ответы в работе_{datetime.now().strftime('%d.%m')}_на_{timenow}.xlsx")
    with pd.ExcelWriter(processed_file_path, engine='openpyxl') as writer:
        df_with_totals.to_excel(writer, sheet_name='СВОД', index=True, startrow=2)
        pivot_prosrok.to_excel(writer, sheet_name='просрочки', index=True, startrow=2)
        pivot_8.to_excel(writer, sheet_name='8-й день', index=True, startrow=2)
        pivot_7.to_excel(writer, sheet_name='7-й день', index=True, startrow=2)
        pivot_6.to_excel(writer, sheet_name='6-й день', index=True, startrow=2)
        pivot_5.to_excel(writer, sheet_name='5-й день', index=True, startrow=2)
        pivot_weekend.to_excel(writer, sheet_name='Выходные статусы', index=True, startrow=2)  # НОВЫЙ ЛИСТ
        main_df.to_excel(writer, sheet_name='Ответы в работе', index=False, startrow=0)
        holidays_df.to_excel(writer, sheet_name='Выходные', index=False, startrow=0)
        prefect.to_excel(writer, sheet_name='Префект просрок', index=False, startrow=0)

    return processed_file_path


def personalizating_table_osn(timenow):
    from openpyxl.utils import get_column_letter  # можно вынести в импорты сверху файла

    file_path = os.path.join(directory, f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')
    wb = load_workbook(file_path)
    ws = wb.worksheets[0]

    header_row = 3
    data_first = 4
    total_row = ws.max_row          # строка "Итог по округу"
    data_last = total_row - 1
    last_col = ws.max_column
    last_col_letter = get_column_letter(last_col)

    # --- стили ---
    light_blue_fill = PatternFill(start_color="5286d1", end_color="5286d1", fill_type="solid")
    pale_blue_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc", fill_type="solid")
    pink_fill = PatternFill(start_color="f7867e", end_color="f7867e", fill_type="solid")
    thin = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def tnr(bold=False, color=None):
        return Font(name='Times New Roman', size=11, bold=bold, color=color)

    # --- заголовок отчёта (строка 2) ---
    ws.merge_cells(f"A2:{last_col_letter}2")
    ws['A2'] = (f'Сводная информация по нарушениям сроков подготовки ответов на сообщения, поступившие на '
                f'централизованный портал "Наш город" по состоянию на {timenow} '
                f'{datetime.now().strftime("%d.%m.%y")} г.')
    ws['A2'].font = tnr(bold=True)
    ws['A2'].alignment = center
    for cell in ws[f'A2:{last_col_letter}2'][0]:
        cell.border = thin_border
    ws.row_dimensions[2].height = 37

    # --- шапка таблицы (строка 3) ---
    for cell in ws[header_row]:
        cell.fill = light_blue_fill
        cell.font = tnr(bold=True)
        cell.border = thin_border
        cell.alignment = center
    ws.row_dimensions[header_row].height = 55

    # --- определяем красные и голубые столбцы по заголовкам ---
    titles = {c.column: str(c.value or "") for c in ws[header_row]}
    idx_5 = next((i for i, t in titles.items() if t.startswith('5 день')), None)
    pink_cols, pale_extra_cols = [], []
    for i, t in titles.items():
        if i == 1:
            continue
        if t in ('Кабинет префекта просрочки', 'Просрочки') or t.startswith(('8 день', '7 день', '6 день', '5 день')):
            pink_cols.append(i)
        elif t.startswith('Вых. дни'):
            if idx_5 is not None and i < idx_5:   # выходные в диапазоне 8–5 дня
                pink_cols.append(i)               # → подсвечиваем красным
            else:
                pale_extra_cols.append(i)         # обычные выходные → голубым
        elif t.startswith(('4 день', '3 день', '2 день', '1 день')):
            pale_extra_cols.append(i)

    # --- данные (строки 4..data_last) ---
    for row in ws.iter_rows(min_row=data_first, max_row=data_last, min_col=1, max_col=last_col):
        for cell in row:
            cell.font = tnr()
            cell.border = thin_border
            cell.alignment = center
            col = cell.column
            if col == 1:                       # район
                cell.fill = pale_blue_fill
                cell.font = tnr(bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col == 2:                     # Всего в работе
                cell.fill = pale_blue_fill
            elif col == 3:                     # Кабинет префекта просрочки
                cell.fill = pale_blue_fill
                cell.font = tnr(bold=True)
            elif col == 4:                     # Всего срочных
                cell.fill = pale_blue_fill
                cell.font = tnr(bold=True, color="800000")
            elif col == 5:                     # Просрочки
                cell.font = tnr(bold=True)
            elif col in pale_extra_cols:       # дни 4–1 и обычные выходные
                cell.fill = pale_blue_fill

    # --- итоговая строка ---
    for cell in ws[total_row]:
        cell.font = tnr(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- условная заливка красным (>0) для срочных столбцов ---
    for col_idx in pink_cols:
        cl = get_column_letter(col_idx)
        rng = f"{cl}{data_first}:{cl}{data_last}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=pink_fill))

    # --- ширина столбцов ---
    ws.column_dimensions['A'].width = 23
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    for col in range(6, last_col + 1):         # столбцы дней и выходных
        ws.column_dimensions[get_column_letter(col)].width = 7.5

    # --- высота строк ---
    for r in range(data_first, total_row + 1):
        ws.row_dimensions[r].height = 14.5

    wb.save(file_path)
    print(f'Formatting applied to the first table in {file_path} successfully.')


def personalizating_table_weekend(timenow):
    """Форматирование листа "Выходные статусы".
    Заголовок в PDF: "Сообщения в выходные дни в разрезе по статусам".
    Структура и раскраска — как у листов "N-й день"; лист = worksheets[6] (7-й по счёту).
    ВЫЗЫВАТЬ ПЕРЕД add_run_delete_and_save_files, иначе в PDF лист будет неокрашенным."""
    file_path = os.path.join(directory, f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')

    wb = load_workbook(file_path)
    ws = wb.worksheets[6]  # 0:СВОД 1:просрочки 2:8д 3:7д 4:6д 5:5д 6:ВЫХОДНЫЕ СТАТУСЫ

    start_row = 3
    max_row = ws.max_row
    max_column = ws.max_column

    header_fill = PatternFill(start_color="5286d1", end_color="5286d1", fill_type="solid")
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc", fill_type="solid")
    red_font = Font(color="FF0000", bold=True)
    bold_font = Font(bold=True)
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Заголовок таблицы (строка 2)
    if max_column > 0:
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = 'Сообщения в выходные дни в разрезе по статусам'
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Шапка таблицы (строка 3)
    header_row = ws[start_row]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Тело таблицы: голубая заливка + красный шрифт по статусам
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for cell in row:
            cell.fill = body_fill
            if cell.column != 1 and cell.column != max_column:
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue
                if row[0].row == max_row:
                    continue
                cell.font = red_font
    wb.save(file_path)
    print(f'Formatting applied to the weekend-status sheet in {file_path} successfully.')


def personalizating_table_prosrok(timenow):
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[1]  # Индексация начинается с 0, поэтому 1 - это второй лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = 'Просроченные сообщения в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the first table in the first sheet in {file_path} successfully.')


def personalizating_table_eight_day(timenow):
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[2]  # Индексация начинается с 0, поэтому 1 - это второй лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = '8-й день в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the first table in the first sheet in {file_path} successfully.')


def personalizating_table_seven_day(timenow):
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[3]  # Индексация начинается с 0, поэтому 3 - это 4 лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = '7-й день в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the first table in the first sheet in {file_path} successfully.')


def personalizating_table_six_day(timenow):
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[4]  # Индексация начинается с 0, поэтому 3 - это 4 лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = '6-й день в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the first table in the first sheet in {file_path} successfully.')


def personalizating_table_five_day(timenow):
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[5]  # Индексация начинается с 0, поэтому 3 - это 4 лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = '5-й день в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the five table in the first sheet in {file_path} successfully.')


def add_run_delete_and_save_files(timenow):
    # Открываем Excel через COM
    file_path = os.path.join(directory,
                             f'Ответы в работе_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx')
    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = True

    wb = excel.Workbooks.Open(os.path.abspath(file_path))

    # Код макроса: добавлен лист 7 (Выходные статусы) — копируется в СВОД и удаляется
    vba_code = """  
Sub CopyTablesToFirstSheet()  
            Dim wsFirst As Worksheet  
            Dim wsSecond As Worksheet  
            Dim wsThird As Worksheet  
            Dim wsFour As Worksheet 
            Dim wsFive As Worksheet 
            Dim wsSix As Worksheet
            Dim wsSeven As Worksheet
            Dim lastRow As Long  
            Dim copyRange As Range  
            ' Установите ссылки на листы  
            Set wsFirst = ThisWorkbook.Worksheets(1)  ' СВОД  
            Set wsSecond = ThisWorkbook.Worksheets(2) ' просрочки  
            Set wsThird = ThisWorkbook.Worksheets(3)  ' 8-й день  
            Set wsFour = ThisWorkbook.Worksheets(4)   ' 7-й день 
            Set wsFive = ThisWorkbook.Worksheets(5)   ' 6-й день
            Set wsSix = ThisWorkbook.Worksheets(6)    ' 5-й день
            Set wsSeven = ThisWorkbook.Worksheets(7)  ' Выходные статусы
            ' Копирование из второго листа (просрочки)  
            lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2  
            Set copyRange = wsSecond.UsedRange  
            copyRange.Copy wsFirst.Cells(lastRow, 1)  
            ' 8-й день  
            lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2  
            Set copyRange = wsThird.UsedRange  
            copyRange.Copy wsFirst.Cells(lastRow, 1)  
            ' 7-й день  
            lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2  
            Set copyRange = wsFour.UsedRange  
            copyRange.Copy wsFirst.Cells(lastRow, 1)  
            ' 6-й день  
            lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2  
            Set copyRange = wsFive.UsedRange  
            copyRange.Copy wsFirst.Cells(lastRow, 1)
            ' 5-й день  
            lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2  
            Set copyRange = wsSix.UsedRange  
            copyRange.Copy wsFirst.Cells(lastRow, 1)
            ' Выходные статусы  
            lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2  
            Set copyRange = wsSeven.UsedRange  
            copyRange.Copy wsFirst.Cells(lastRow, 1)
            ' Очистка буфера обмена  
            Application.CutCopyMode = False  
            ' Удаление скопированных листов  
            Application.DisplayAlerts = False  
            wsSecond.Delete  
            wsThird.Delete  
            wsFour.Delete
            wsFive.Delete
            wsSix.Delete
            wsSeven.Delete
            Application.DisplayAlerts = True  
        End Sub  

        Sub DeleteFirstSheet()  
            Dim wsFirst As Worksheet  
            Set wsFirst = ThisWorkbook.Worksheets(1)  
            Application.DisplayAlerts = False  
            wsFirst.Delete  
            Application.DisplayAlerts = True  
        End Sub
    """

    # Добавляем модуль в книгу и вставляем код
    vba_module = wb.VBProject.VBComponents.Add(1)  # 1 - стандартный модуль
    vba_module.Name = 'MyMacroModule'
    vba_module.CodeModule.AddFromString(vba_code)

    # Выполнение макроса
    excel.Application.Run('MyMacroModule.CopyTablesToFirstSheet')

    # Сохранение первого листа как PDF
    pdf_file_name = f'{datetime.now().strftime("%d.%m")}_на_{timenow}.pdf'
    pdf_path = os.path.join(os.path.dirname(file_path), pdf_file_name)
    wsFirst = wb.Worksheets(1)

    wsFirst.PageSetup.FitToPagesWide = 1
    wsFirst.PageSetup.FitToPagesTall = 1
    wsFirst.PageSetup.Zoom = False
    wsFirst.PageSetup.LeftMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.RightMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.TopMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.BottomMargin = excel.Application.CentimetersToPoints(0.5)
    wb.Save()
    try:
        if os.path.exists(pdf_path):
            print(f"Файл {pdf_path} существует. Удаление...")
            os.remove(pdf_path)
            print("Файл успешно удален.")
        print(f"Сохранение файла в {pdf_path}...")
        wsFirst.ExportAsFixedFormat(0, pdf_path)  # 0 - xlTypePDF
        print(f"PDF успешно создан: {pdf_path}")
    except Exception as e:
        print(f"Ошибка при сохранении файла: {e}")

    first_sheet_file_name = f'CВОД_{datetime.now().strftime("%d.%m")}_на_{timenow}.xlsx'
    first_sheet_file_path = os.path.join(directory, first_sheet_file_name)
    print(first_sheet_file_path)

    wsFirst.Copy()
    print(1)
    wb_first_sheet = excel.ActiveWorkbook

    try:
        if os.path.exists(first_sheet_file_path):
            print(f"Файл {first_sheet_file_path} существует. Удаление...")
            os.remove(first_sheet_file_path)
            print("Файл успешно удален.")
        print(f"Сохранение файла в {first_sheet_file_path}...")
        wb_first_sheet.SaveAs(first_sheet_file_path, FileFormat=51)
        print("Файл успешно сохранен.")
    except Exception as e:
        print(f"Ошибка при сохранении файла: {e}")
    finally:
        wb_first_sheet.Close()
        print(first_sheet_file_path)

    # Удаление первого листа из основного файла через макрос
    excel.Application.Run('MyMacroModule.DeleteFirstSheet')

    # Авторазмер колонок на оставшихся листах
    for sheet in wb.Worksheets:
        sheet.Cells.EntireColumn.AutoFit()

    wb.Save()
    wb.Close()
    excel.Quit()
    return pdf_path, first_sheet_file_path, file_path
